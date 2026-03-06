"""Генерация Excel: фиксированный шаблон + шаблонные колонки + динамические колонки."""

from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Служебные ключи — не показываем как колонки данных
_INTERNAL = frozenset({
    "filename", "page", "parse_error", "raw_text", "error",
    "проблемы", "уверенность", "extracted",
    "_страница_без_данных", "_предупреждения",
})

# Фиксированный шаблон колонок — 15 колонок по «Названия столбцов.xlsx».
# Покупатель (контрагент2) идёт ПЕРЕД продавцом (контрагент1).
# (заголовок_в_Excel, ключ_в_JSON или спец-значение)
TEMPLATE_COLUMNS = [
    ("Путь к файлу", "_link"),
    ("Имя файла", "_filename"),
    ("Тип документа", "тип_документа"),
    ("Номер документа", "номер_документа"),
    ("Дата документа", "дата_документа"),
    ("Покупатель", "контрагент2"),
    ("ИНН", "контрагент2_инн"),
    ("КПП", "контрагент2_кпп"),
    ("Адрес", "контрагент2_адрес"),
    ("Продавец", "контрагент1"),
    ("ИНН", "контрагент1_инн"),
    ("КПП", "контрагент1_кпп"),
    ("Адрес", "контрагент1_адрес"),
    ("НДС", "ндс"),
    ("Сумма с НДС", "сумма_с_ндс"),
]


def _flatten(data: dict, prefix: str = "") -> dict:
    """Развернуть вложенные dict в плоскую структуру с составными ключами."""
    flat: dict[str, str] = {}
    for key, value in data.items():
        if key in _INTERNAL:
            continue
        full_key = f"{prefix} → {key}" if prefix else key

        if isinstance(value, dict):
            flat.update(_flatten(value, full_key))
        elif isinstance(value, list):
            if value and isinstance(value[0], dict):
                parts = []
                for item in value:
                    parts.append(", ".join(
                        f"{v}" for v in item.values() if v is not None
                    ))
                flat[full_key] = "; ".join(parts)
            elif value:
                flat[full_key] = "; ".join(str(v) for v in value if v)
        else:
            flat[full_key] = value
    return flat


def _make_styles():
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    link_font = Font(color="0563C1", underline="single", size=11)
    return header_font, header_fill, thin_border, link_font


def _auto_width(ws, num_cols, num_rows):
    for col_idx in range(1, num_cols + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, num_rows + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def _generate_template_excel(
    parsed_results: list[dict],
    batch_id: str,
    base_url: str,
) -> bytes:
    """Excel по фиксированному шаблону (15 колонок)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font, header_fill, thin_border, link_font = _make_styles()

    # Заголовки
    for col, (header, _) in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Данные
    for idx, result in enumerate(parsed_results, 1):
        row_num = idx + 1
        filename = result.get("filename", "")

        for col, (header, json_key) in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

            if json_key == "_link":
                # Путь к файлу — гиперссылка на PDF
                if base_url and batch_id and filename:
                    url = f"{base_url}/api/files/{batch_id}/{quote(filename)}"
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.value = filename
            elif json_key == "_filename":
                cell.value = filename
            else:
                value = result.get(json_key)
                if value is not None and value != "" and value != "null":
                    cell.value = value if not isinstance(value, (int, float)) else value
                else:
                    cell.value = ""

    # Авто-ширина
    total_rows = len(parsed_results) + 2
    _auto_width(ws, len(TEMPLATE_COLUMNS), total_rows)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_from_template(
    parsed_results: list[dict],
    batch_id: str,
    base_url: str,
    template: dict,
) -> bytes:
    """Excel с колонками из шаблона документа."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font, header_fill, thin_border, link_font = _make_styles()

    fields = template.get("fields", [])

    # Заголовки: Путь к файлу + Имя файла + поля из шаблона
    headers = [("Путь к файлу", "_link"), ("Имя файла", "_filename")]
    headers += [(f["label"], f["key"]) for f in fields]

    for col, (header, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Данные
    for idx, result in enumerate(parsed_results, 1):
        row_num = idx + 1
        filename = result.get("filename", "")

        for col, (header, json_key) in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

            if json_key == "_link":
                if base_url and batch_id and filename:
                    url = f"{base_url}/api/files/{batch_id}/{quote(filename)}"
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = link_font
                else:
                    cell.value = filename
            elif json_key == "_filename":
                cell.value = filename
            else:
                value = result.get(json_key)
                if value is not None and value != "" and value != "null":
                    cell.value = value if not isinstance(value, (int, float)) else value
                else:
                    cell.value = ""

    total_rows = len(parsed_results) + 2
    _auto_width(ws, len(headers), total_rows)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_dynamic_excel(
    parsed_results: list[dict],
    batch_id: str,
    base_url: str,
) -> bytes:
    """Excel с динамическими колонками (для гибкого режима с custom_prompt)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font, header_fill, thin_border, link_font = _make_styles()

    # Развернуть и собрать все ключи
    flat_rows: list[tuple] = []
    all_keys: list[str] = []

    for result in parsed_results:
        fields = result.get("extracted", result)
        flat = _flatten(fields)
        flat_rows.append((result, flat))
        for key in flat:
            if key not in all_keys:
                all_keys.append(key)

    # Заголовки
    headers = ["Документ (PDF)"] + all_keys

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Данные
    for idx, (result, flat) in enumerate(flat_rows, 1):
        row_num = idx + 1
        filename = result.get("filename", "")

        cell = ws.cell(row=row_num, column=1, value=filename)
        cell.border = thin_border
        if base_url and batch_id and filename:
            url = f"{base_url}/api/files/{batch_id}/{quote(filename)}"
            cell.hyperlink = url
            cell.font = link_font

        for col_offset, key in enumerate(all_keys):
            value = flat.get(key)
            cell = ws.cell(
                row=row_num,
                column=col_offset + 2,
                value=str(value) if value is not None else "",
            )
            cell.border = thin_border

    # Авто-ширина
    total_rows = len(parsed_results) + 2
    _auto_width(ws, len(headers), total_rows)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel(
    parsed_results: list[dict],
    batch_id: str = "",
    base_url: str = "",
    template_id: str = "",
    is_template_mode: bool = False,
) -> bytes:
    """Генерация Excel.

    Роутинг:
    - custom_prompt (нет template_id, не template_mode) → динамические колонки
    - template_id указан → колонки из fields шаблона
    - универсальный режим (нет template_id) → фиксированные 15 колонок
    """
    if not template_id and not is_template_mode:
        # Custom prompt mode — произвольные колонки
        return _generate_dynamic_excel(parsed_results, batch_id, base_url)

    # Если указан шаблон — используем его fields для колонок
    if template_id:
        from app.template_registry import get_template
        template = get_template(template_id)
        if template and template.get("fields"):
            return _generate_from_template(parsed_results, batch_id, base_url, template)

    # Универсальный режим — фиксированные 15 колонок
    return _generate_template_excel(parsed_results, batch_id, base_url)
